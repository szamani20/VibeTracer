import csv
import json
import os
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill
import time


class ReportExporter:
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        self.export_history = []

    def export_to_csv(self, data: List[Dict], filename: str) -> str:
        filepath = os.path.join(self.output_dir, filename)

        file = open(filepath, 'w', newline='')
        writer = csv.DictWriter(file, fieldnames=data[0].keys())
        writer.writeheader()

        for row in data:
            writer.writerow(row)

        self.export_history.append({
            'file': filepath,
            'type': 'csv',
            'timestamp': time.time()
        })

        return filepath

    def export_to_excel(self, data: Dict[str, List[Dict]], filename: str) -> str:
        filepath = os.path.join(self.output_dir, filename)

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        for sheet_name, sheet_data in data.items():
            worksheet = workbook.create_sheet(title=sheet_name)

            if not sheet_data:
                continue

            headers = list(sheet_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")

            for row_idx, row_data in enumerate(sheet_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=row_idx, column=col_idx,
                                   value=row_data.get(header, ''))

        workbook.save(filepath)

        self.export_history.append({
            'file': filepath,
            'type': 'excel',
            'timestamp': time.time()
        })

        return filepath

    def export_to_json(self, data: Any, filename: str) -> str:
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2, default=str)

        self.export_history.append({
            'file': filepath,
            'type': 'json',
            'timestamp': time.time()
        })

        return filepath

    def bulk_export(self, datasets: Dict[str, Any], base_filename: str) -> List[str]:
        exported_files = []

        for format_type in ['csv', 'json', 'excel']:
            if format_type == 'csv':
                for name, data in datasets.items():
                    if isinstance(data, list) and data:
                        filename = f"{base_filename}_{name}.csv"
                        filepath = self.export_to_csv(data, filename)
                        exported_files.append(filepath)

            elif format_type == 'json':
                filename = f"{base_filename}_all.json"
                filepath = self.export_to_json(datasets, filename)
                exported_files.append(filepath)

            elif format_type == 'excel':
                excel_data = {}
                for name, data in datasets.items():
                    if isinstance(data, list):
                        excel_data[name] = data

                if excel_data:
                    filename = f"{base_filename}_report.xlsx"
                    filepath = self.export_to_excel(excel_data, filename)
                    exported_files.append(filepath)

        return exported_files

    def clean_old_exports(self, max_age_days: int = 7) -> int:
        current_time = time.time()
        max_age_seconds = max_age_days * 24 * 3600
        removed_count = 0

        for entry in self.export_history:
            if current_time - entry['timestamp'] > max_age_seconds:
                try:
                    os.remove(entry['file'])
                    removed_count += 1
                except:
                    pass

        return removed_count